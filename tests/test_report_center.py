"""报表中心验收测试。"""

from __future__ import annotations

import csv
from datetime import date

from openpyxl import load_workbook

from mym2.db.models.account import Account
from mym2.db.models.budget import BudgetLine, BudgetPeriod
from mym2.db.models.category import Category
from mym2.db.models.transaction import Transaction
from mym2.domain.enums import AccountType
from mym2.services.report_service import ReportFilter, ReportService
from mym2.ui.pages.reports_page import ReportsPage


def _seed_report_data(session):
    cash = Account(
        name='现金',
        type=AccountType.CASH.value,
        is_enabled=True,
        current_balance_minor=120000,
    )
    receivable = Account(
        name='张三应收',
        type=AccountType.RECEIVABLE.value,
        is_enabled=True,
        current_balance_minor=3000,
    )
    investment = Account(
        name='历史投资资产快照',
        type=AccountType.INVESTMENT_SNAPSHOT.value,
        is_enabled=True,
        is_editable=False,
        is_locked=True,
        current_balance_minor=500000,
    )
    credit = Account(
        name='信用卡',
        type=AccountType.CREDIT_CARD.value,
        is_enabled=True,
        current_balance_minor=20000,
    )
    food = Category(name='餐饮', type='expense', is_enabled=True)
    salary = Category(name='工资', type='income', is_enabled=True)
    session.add_all([cash, receivable, investment, credit, food, salary])
    session.flush()

    today = date.today()
    rows = [
        Transaction(
            transaction_date=today,
            type='expense',
            account_out_id=cash.id,
            category_id=food.id,
            amount_minor=1234,
            note='早餐',
        ),
        Transaction(
            transaction_date=today,
            type='income',
            account_out_id=cash.id,
            account_in_id=cash.id,
            category_id=salary.id,
            amount_minor=678900,
            note='工资收入',
        ),
        Transaction(
            transaction_date=today,
            type='balance_adjustment',
            account_out_id=cash.id,
            amount_minor=999999,
            note='不计入日常',
        ),
        Transaction(
            transaction_date=today,
            type='historical_investment_settlement',
            account_out_id=investment.id,
            amount_minor=888888,
            note='历史估值',
            is_locked=True,
        ),
        Transaction(
            transaction_date=today,
            type='receivable_advance',
            account_out_id=cash.id,
            account_in_id=receivable.id,
            amount_minor=5000,
            note='代付',
        ),
        Transaction(
            transaction_date=today,
            type='receivable_repayment',
            account_out_id=receivable.id,
            account_in_id=cash.id,
            amount_minor=2000,
            note='=还款公式防护',
        ),
    ]
    session.add_all(rows)
    period = BudgetPeriod(year=today.year, month=today.month)
    session.add(period)
    session.flush()
    session.add(
        BudgetLine(
            budget_period_id=period.id,
            category_id=food.id,
            type='expense',
            amount_minor=20000,
        )
    )
    session.commit()
    return cash, receivable, investment, credit, food, salary


def _current_filter() -> ReportFilter:
    today = date.today()
    return ReportFilter(
        start_date=date(today.year, today.month, 1),
        end_date=today,
    )


def test_daily_reports_exclude_adjustment_historical_and_receivable_principal(session):
    svc = ReportService()
    _seed_report_data(session)

    category = svc.query(session, 'category_income_expense', _current_filter())
    monthly = svc.query(session, 'monthly_income_expense', _current_filter())

    assert category.summary['expense_minor'] == 1234
    assert category.summary['income_minor'] == 678900
    assert monthly.summary['expense_minor'] == 1234
    assert monthly.summary['income_minor'] == 678900
    assert '历史估值' not in {row['category_name'] for row in category.rows}


def test_dashboard_and_report_totals_are_consistent(session):
    svc = ReportService()
    _seed_report_data(session)

    dashboard = svc.get_dashboard_data(session)
    monthly = svc.query(session, 'monthly_income_expense', _current_filter())
    balance = svc.query(session, 'asset_liability', _current_filter())

    assert dashboard.current_month_income_minor == monthly.summary['income_minor']
    assert dashboard.current_month_expense_minor == monthly.summary['expense_minor']
    assert dashboard.total_assets_minor == balance.summary['total_assets_minor']
    assert dashboard.net_worth_minor == balance.summary['net_worth_minor']


def test_asset_report_includes_investment_snapshot_without_functional_stock_rows(session):
    svc = ReportService()
    _seed_report_data(session)

    result = svc.query(session, 'asset_liability', _current_filter())
    names = [row['account_name'] for row in result.rows]
    types = [row['account_type'] for row in result.rows]

    assert '历史投资资产快照' in names
    assert '历史投资资产快照' in types
    assert result.summary['total_assets_minor'] == 623000
    assert all('持仓' not in str(row) and '行情' not in str(row) for row in result.rows)


def test_budget_execution_uses_same_expense_scope(session):
    svc = ReportService()
    _seed_report_data(session)

    result = svc.query(session, 'budget_execution', _current_filter())

    assert result.summary['budget_total_minor'] == 20000
    assert result.summary['actual_total_minor'] == 1234
    assert result.rows[0]['remaining_minor'] == 18766


def test_csv_and_excel_export_dates_amounts_and_utf8(session, tmp_path):
    svc = ReportService()
    _seed_report_data(session)
    result = svc.query(session, 'transaction_detail', _current_filter())

    csv_path = tmp_path / '报表.csv'
    xlsx_path = tmp_path / '报表.xlsx'
    svc.export_csv(result, csv_path)
    svc.export_excel(result, xlsx_path)

    raw = csv_path.read_bytes()
    assert raw.startswith(b'\xef\xbb\xbf')
    with csv_path.open('r', encoding='utf-8-sig', newline='') as fh:
        rows = list(csv.reader(fh))
    assert '日期' in rows[0]
    assert '金额' in rows[0]
    assert any('6789.00' in row for row in rows)
    assert any('工资收入' in row for row in rows)
    assert any("'=还款公式防护" in row for row in rows)

    wb = load_workbook(xlsx_path)
    ws = wb.active
    assert ws['A1'].value == '日期'
    values = [[cell.value for cell in row] for row in ws.iter_rows()]
    assert any('6789.00' in row for row in values)
    assert any('工资收入' in row for row in values)


def test_pdf_print_html_is_offline_and_pdf_is_generated(session, tmp_path, qapp):
    svc = ReportService()
    _seed_report_data(session)
    result = svc.query(session, 'category_income_expense', _current_filter())

    html = svc.build_print_html(result)
    assert 'http://' not in html
    assert 'https://' not in html
    assert 'cdn' not in html.lower()

    pdf_path = tmp_path / 'report.pdf'
    svc.export_pdf(result, pdf_path)
    assert pdf_path.read_bytes().startswith(b'%PDF')
    assert pdf_path.stat().st_size > 1000


def test_reports_page_constructs_and_has_export_controls(qapp):
    page = ReportsPage()
    assert page._table is not None
    assert page._csv_btn is not None
    assert page._excel_btn is not None
    assert page._pdf_btn is not None
    page.deleteLater()

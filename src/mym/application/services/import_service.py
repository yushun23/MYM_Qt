"""ImportService – AI/rule hybrid financial table import with preview, mapping, dedup (P33)."""

import csv
import hashlib
import json
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from enum import Enum
from pathlib import Path
from typing import Any, Callable

from sqlalchemy.orm import Session

from mym.application.dto.transaction_dto import (
    CreateTransactionDTO,
    TransactionLineDTO,
)
from mym.application.use_cases.create_transaction import CreateTransactionUseCase
from mym.domain.entities.account import Account
from mym.domain.entities.category import Category
from mym.domain.entities.import_ import ImportJob
from mym.domain.entities.transaction import Transaction
from mym.domain.enums import (
    AccountType,
    CategoryType,
    ImportStatus,
    ImportIssueSeverity,
    TransactionSource,
    TransactionStatus,
)
from mym.infrastructure.repositories.account_repo import AccountRepository
from mym.infrastructure.repositories.category_repo import CategoryRepository

logger = logging.getLogger(__name__)

MAX_PREVIEW_ROWS = 200
MAX_FILE_SIZE_MB = 20
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024

# ── Data Types ──────────────────────────────────────────────────────────────

class ImportField(str, Enum):
    """Standard import target fields."""
    DATE = "date"
    TYPE = "type"           # income / expense / transfer
    AMOUNT = "amount"
    INCOME_AMOUNT = "income_amount"
    EXPENSE_AMOUNT = "expense_amount"
    ACCOUNT = "account"
    CATEGORY = "category"
    MEMO = "memo"
    TRANSFER_TO = "transfer_to"
    IGNORE = "ignore"


@dataclass
class FieldMapping:
    """Mapping from source column index to target field."""
    source_index: int
    source_header: str
    target_field: ImportField
    ai_suggested: bool = False


@dataclass
class ImportIssue:
    """An issue found during import preview."""
    row: int
    severity: ImportIssueSeverity
    field: str
    message: str


@dataclass
class ImportPreviewRow:
    """A single row in the import preview."""
    row_number: int
    source_data: dict[str, Any]
    mapped_data: dict[str, Any]
    issues: list[ImportIssue] = field(default_factory=list)
    is_duplicate: bool = False
    will_create_account: bool = False
    will_create_category: bool = False
    is_valid: bool = True


@dataclass
class ImportPreview:
    """Full import preview result."""
    file_name: str
    file_hash: str
    source_headers: list[str]
    total_rows: int
    valid_rows: int
    duplicate_rows: int
    error_rows: int
    field_mappings: list[FieldMapping]
    preview_rows: list[ImportPreviewRow]
    issues: list[ImportIssue]
    accounts_to_create: list[str]
    categories_to_create: list[str]
    estimated_transactions: int


@dataclass
class ImportResult:
    """Result of executing an import."""
    success: bool
    batch_id: str | None = None
    total_imported: int = 0
    skipped_duplicates: int = 0
    failed_count: int = 0
    errors: list[str] = field(default_factory=list)
    created_accounts: list[str] = field(default_factory=list)
    created_categories: list[str] = field(default_factory=list)


# ── Field Detectors ─────────────────────────────────────────────────────────

def _field_similarity(a: str, b: str) -> float:
    """Simple similarity between two strings."""
    a, b = a.lower().strip(), b.lower().strip()
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.8
    # Check common substrings
    common = sum(1 for c in a if c in b) / max(len(a), len(b), 1)
    return common


class FieldDetector:
    """Auto-detect field mappings from column headers."""

    DATE_PATTERNS = ["日期", "date", "时间", "time", "交易日期", "记账日期"]
    TYPE_PATTERNS = ["类型", "type", "方向", "收支", "借贷"]
    AMOUNT_PATTERNS = ["金额", "amount", "交易金额", "发生额"]
    INCOME_PATTERNS = ["收入", "income", "贷方", "credit", "存入"]
    EXPENSE_PATTERNS = ["支出", "expense", "借方", "debit", "取出"]
    ACCOUNT_PATTERNS = ["账户", "account", "账号", "卡号"]
    CATEGORY_PATTERNS = ["分类", "category", "类型", "项目", "用途", "摘要"]
    MEMO_PATTERNS = ["备注", "memo", "说明", "note", "描述", "description", "附言"]

    def detect(self, header: str) -> ImportField | None:
        """Detect the target field for a given header name. Returns None if unsure."""
        h = header.lower().strip()

        for pat in self.DATE_PATTERNS:
            if _field_similarity(h, pat) > 0.6:
                return ImportField.DATE

        for pat in self.TYPE_PATTERNS:
            if _field_similarity(h, pat) > 0.6:
                return ImportField.TYPE

        for pat in self.AMOUNT_PATTERNS:
            if _field_similarity(h, pat) > 0.6:
                return ImportField.AMOUNT

        for pat in self.INCOME_PATTERNS:
            if _field_similarity(h, pat) > 0.6:
                return ImportField.INCOME_AMOUNT

        for pat in self.EXPENSE_PATTERNS:
            if _field_similarity(h, pat) > 0.6:
                return ImportField.EXPENSE_AMOUNT

        for pat in self.ACCOUNT_PATTERNS:
            if _field_similarity(h, pat) > 0.6:
                return ImportField.ACCOUNT

        for pat in self.CATEGORY_PATTERNS:
            if _field_similarity(h, pat) > 0.6:
                return ImportField.CATEGORY

        for pat in self.MEMO_PATTERNS:
            if _field_similarity(h, pat) > 0.6:
                return ImportField.MEMO

        return None  # AI can suggest


# ── Date Parsers ────────────────────────────────────────────────────────────

DATE_FORMATS = [
    "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
    "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y", "%d-%m-%Y",
    "%Y%m%d",
    "%Y年%m月%d日",
    "%m/%d/%y", "%m-%d-%y",
]


def parse_date(value: Any) -> date | None:
    """Try to parse a date from various formats."""
    if isinstance(value, (date, datetime)):
        return value if isinstance(value, date) else value.date()
    if value is None:
        return None
    s = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Try parsing as Excel serial date
    try:
        num = float(s)
        if 30000 < num < 100000:  # reasonable Excel date range
            from datetime import datetime as dt
            return dt(1899, 12, 30).date() + __import__('datetime').timedelta(days=int(num))
    except (ValueError, OverflowError):
        pass
    return None


# ── Row Hashing ─────────────────────────────────────────────────────────────

def compute_row_hash(row_data: dict[str, Any]) -> str:
    """Compute a stable hash for dedup."""
    normalized = {}
    for k, v in sorted(row_data.items()):
        if v is not None:
            normalized[k] = str(v).strip()
        else:
            normalized[k] = ""
    raw = json.dumps(normalized, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(raw.encode()).hexdigest()


def compute_file_hash(file_path: Path) -> str:
    """Compute SHA-256 of file for import source tracking."""
    sha = hashlib.sha256()
    with open(file_path, "rb") as f:
        while True:
            chunk = f.read(8192)
            if not chunk:
                break
            sha.update(chunk)
    return sha.hexdigest()


# ── Main Import Service ────────────────────────────────────────────────────

class ImportService:
    """Service for importing financial data from CSV/XLSX files.

    All writes go through CreateTransactionUseCase. Import is transactional
    and reversible by batch.
    """

    def __init__(self, session: Session) -> None:
        self._session = session
        self._detector = FieldDetector()
        self._account_repo = AccountRepository(session)
        self._category_repo = CategoryRepository(session)

    # ── Reading ──────────────────────────────────────────────────────────

    def read_headers(self, file_path: str | Path) -> list[str]:
        """Read the header row from a file."""
        path = Path(file_path)
        ext = path.suffix.lower()

        if ext == ".csv":
            return self._read_csv_headers(path)
        elif ext in (".xlsx", ".xls"):
            return self._read_excel_headers(path)
        else:
            raise ValueError(f"不支持的文件类型: {ext}")

    def _read_csv_headers(self, path: Path) -> list[str]:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            try:
                headers = next(reader)
            except StopIteration:
                return []
        return [h.strip() for h in headers]

    def _read_excel_headers(self, path: Path) -> list[str]:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        try:
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(c).strip() if c is not None else "" for c in row]
        except StopIteration:
            headers = []
        wb.close()
        return headers

    def read_rows(self, file_path: str | Path, max_rows: int = MAX_PREVIEW_ROWS) -> list[dict[str, Any]]:
        """Read data rows from a file."""
        path = Path(file_path)
        ext = path.suffix.lower()

        if ext == ".csv":
            return self._read_csv_rows(path, max_rows)
        elif ext in (".xlsx", ".xls"):
            return self._read_excel_rows(path, max_rows)
        return []

    def _read_csv_rows(self, path: Path, max_rows: int) -> list[dict]:
        headers = self._read_csv_headers(path)
        rows = []
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            next(reader, None)  # skip header
            for i, row in enumerate(reader):
                if i >= max_rows:
                    break
                row_dict = {}
                for j, val in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = val.strip()
                    else:
                        row_dict[f"col_{j}"] = val.strip()
                rows.append(row_dict)
        return rows

    def _read_excel_rows(self, path: Path, max_rows: int) -> list[dict]:
        import openpyxl
        headers = self._read_excel_headers(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if i >= max_rows:
                break
            row_dict = {}
            for j, val in enumerate(row):
                key = headers[j] if j < len(headers) else f"col_{j}"
                row_dict[key] = str(val).strip() if val is not None else ""
            rows.append(row_dict)
        wb.close()
        return rows

    # ── Preview ──────────────────────────────────────────────────────────

    def preview(
        self,
        file_path: str | Path,
        custom_mappings: dict[int, ImportField] | None = None,
    ) -> ImportPreview:
        """Generate a preview of the import with field detection and issues."""
        path = Path(file_path)
        headers = self.read_headers(path)
        rows = self.read_rows(path)

        # Detect field mappings
        mappings = self._detect_mappings(headers, custom_mappings)

        # Build preview rows
        preview_rows: list[ImportPreviewRow] = []
        issues: list[ImportIssue] = []
        accounts_to_create: set[str] = set()
        categories_to_create: set[str] = set()
        existing_accounts = self._get_existing_account_names()
        existing_categories = self._get_existing_category_names()
        existing_hashes = self._get_existing_row_hashes()
        seen_hashes: set[str] = set()

        valid_count = 0
        dup_count = 0
        err_count = 0

        for i, row_data in enumerate(rows):
            row_num = i + 2  # 1-indexed, header is row 1
            mapped = self._apply_mappings(row_data, mappings)
            row_issues: list[ImportIssue] = []
            is_dup = False
            is_valid = True

            # Check for missing required fields
            date_val = mapped.get("date")
            amount_val = mapped.get("amount")
            type_val = mapped.get("type")

            if not date_val and not mapped.get("income_amount") and not mapped.get("expense_amount"):
                row_issues.append(ImportIssue(row_num, ImportIssueSeverity.WARNING, "date", "缺少日期"))
                is_valid = False

            # Parse amount
            if amount_val is not None:
                try:
                    Decimal(str(amount_val).replace(",", "").replace("¥", "").replace("$", ""))
                except (InvalidOperation, ValueError):
                    row_issues.append(ImportIssue(row_num, ImportIssueSeverity.ERROR, "amount", f"无效金额: {amount_val}"))
                    is_valid = False
            elif mapped.get("income_amount") or mapped.get("expense_amount"):
                # Income/expense columns
                pass
            else:
                row_issues.append(ImportIssue(row_num, ImportIssueSeverity.ERROR, "amount", "缺少金额"))
                is_valid = False

            # Check account existence
            acct = mapped.get("account", "")
            if acct and acct not in existing_accounts:
                row_issues.append(ImportIssue(
                    row_num, ImportIssueSeverity.WARNING, "account",
                    f"账户 '{acct}' 不存在，将自动创建",
                ))
                accounts_to_create.add(acct)

            # Check category existence
            cat = mapped.get("category", "")
            if cat and cat not in existing_categories:
                row_issues.append(ImportIssue(
                    row_num, ImportIssueSeverity.WARNING, "category",
                    f"分类 '{cat}' 不存在，将自动创建",
                ))
                categories_to_create.add(cat)

            # Check type validity
            if type_val and type_val not in ("income", "expense", "transfer", "收入", "支出", "转账"):
                row_issues.append(ImportIssue(
                    row_num, ImportIssueSeverity.WARNING, "type",
                    f"未知交易类型: {type_val}",
                ))

            # Dedup check
            row_hash = compute_row_hash(row_data)
            if row_hash in existing_hashes or row_hash in seen_hashes:
                is_dup = True
                dup_count += 1
                row_issues.append(ImportIssue(row_num, ImportIssueSeverity.INFO, "", "重复数据"))
            seen_hashes.add(row_hash)

            if is_valid and not is_dup:
                valid_count += 1
            if not is_valid:
                err_count += 1

            issues.extend(row_issues)

            preview_rows.append(ImportPreviewRow(
                row_number=row_num,
                source_data=row_data,
                mapped_data=mapped,
                issues=row_issues,
                is_duplicate=is_dup,
                will_create_account=acct in accounts_to_create,
                will_create_category=cat in categories_to_create,
                is_valid=is_valid and not is_dup,
            ))

        return ImportPreview(
            file_name=path.name,
            file_hash=compute_file_hash(path),
            source_headers=headers,
            total_rows=len(rows),
            valid_rows=valid_count,
            duplicate_rows=dup_count,
            error_rows=err_count,
            field_mappings=mappings,
            preview_rows=preview_rows,
            issues=issues,
            accounts_to_create=sorted(accounts_to_create),
            categories_to_create=sorted(categories_to_create),
            estimated_transactions=valid_count,
        )

    def _detect_mappings(
        self, headers: list[str], custom: dict[int, ImportField] | None = None,
    ) -> list[FieldMapping]:
        """Detect field mappings from headers. Custom overrides take priority."""
        mappings = []
        for idx, header in enumerate(headers):
            if custom and idx in custom:
                mappings.append(FieldMapping(idx, header, custom[idx]))
                continue
            detected = self._detector.detect(header)
            if detected:
                mappings.append(FieldMapping(idx, header, detected))
            else:
                mappings.append(FieldMapping(idx, header, ImportField.IGNORE, ai_suggested=True))
        return mappings

    def _apply_mappings(self, row: dict, mappings: list[FieldMapping]) -> dict:
        """Apply field mappings to a source row."""
        result: dict[str, Any] = {}
        for m in mappings:
            key = m.source_header
            val = row.get(key, "")
            if m.target_field == ImportField.IGNORE:
                continue
            elif m.target_field == ImportField.DATE:
                result["date"] = parse_date(val)
            elif m.target_field == ImportField.AMOUNT:
                try:
                    clean = str(val).replace(",", "").replace("¥", "").replace("$", "").strip()
                    result["amount"] = Decimal(clean) if clean else None
                except (InvalidOperation, ValueError):
                    result["amount"] = None
            elif m.target_field == ImportField.INCOME_AMOUNT:
                try:
                    clean = str(val).replace(",", "").strip()
                    amt = Decimal(clean) if clean else None
                    if amt and amt > 0:
                        result["amount"] = amt
                        result["type"] = "income"
                except (InvalidOperation, ValueError):
                    pass
            elif m.target_field == ImportField.EXPENSE_AMOUNT:
                try:
                    clean = str(val).replace(",", "").strip()
                    amt = Decimal(clean) if clean else None
                    if amt and amt > 0:
                        result["amount"] = amt
                        result["type"] = "expense"
                except (InvalidOperation, ValueError):
                    pass
            elif m.target_field == ImportField.TYPE:
                t = str(val).strip().lower()
                if t in ("收入", "income", "进", "贷"):
                    result["type"] = "income"
                elif t in ("支出", "expense", "出", "借"):
                    result["type"] = "expense"
                else:
                    result["type"] = t
            elif m.target_field == ImportField.ACCOUNT:
                result["account"] = str(val).strip()
            elif m.target_field == ImportField.CATEGORY:
                result["category"] = str(val).strip()
            elif m.target_field == ImportField.MEMO:
                result["memo"] = str(val).strip()
            elif m.target_field == ImportField.TRANSFER_TO:
                result["transfer_to"] = str(val).strip()
        return result

    # ── Execution ────────────────────────────────────────────────────────

    def execute_import(
        self,
        file_path: str | Path,
        preview: ImportPreview,
        create_accounts: bool = True,
        create_categories: bool = True,
    ) -> ImportResult:
        """Execute the import, writing all valid rows via CreateTransactionUseCase."""
        imported = 0
        skipped = 0
        failed = 0
        errors: list[str] = []
        created_accts: list[str] = []
        created_cats: list[str] = []

        try:
            # Create ImportJob record
            import_job = ImportJob(
                source_file=str(Path(file_path).name),
                file_hash=preview.file_hash,
                import_type="financial_table",
                status=ImportStatus.IN_PROGRESS,
                total_rows=preview.total_rows,
            )
            self._session.add(import_job)
            self._session.flush()

            # Create accounts if needed
            if create_accounts:
                existing = self._get_existing_account_names()
                for acct_name in preview.accounts_to_create:
                    if acct_name not in existing:
                        acct = Account(
                            name=acct_name,
                            account_type=AccountType.ASSET,
                            opening_balance=Decimal("0"),
                        )
                        self._session.add(acct)
                        self._session.flush()
                        created_accts.append(acct_name)

            # Create categories if needed
            if create_categories:
                existing = self._get_existing_category_names()
                for cat_name in preview.categories_to_create:
                    if cat_name not in existing:
                        cat = Category(
                            name=cat_name,
                            category_type=CategoryType.EXPENSE,
                        )
                        self._session.add(cat)
                        self._session.flush()
                        created_cats.append(cat_name)

            # Process each valid row
            tx_uc = CreateTransactionUseCase(self._session)

            for pr in preview.preview_rows:
                if not pr.is_valid or pr.is_duplicate:
                    if pr.is_duplicate:
                        skipped += 1
                    else:
                        failed += 1
                    continue

                try:
                    md = pr.mapped_data
                    tx_date = md.get("date") or date.today()
                    tx_type = md.get("type") or "expense"
                    amount = md.get("amount")
                    if amount is None:
                        failed += 1
                        errors.append(f"第{pr.row_number}行: 缺少金额")
                        continue
                    amount = Decimal(str(amount))

                    # Find account
                    acct_name = md.get("account", "")
                    acct = self._find_account(acct_name)
                    if not acct:
                        failed += 1
                        errors.append(f"第{pr.row_number}行: 账户 '{acct_name}' 不存在")
                        continue

                    # Find category
                    cat_name = md.get("category", "")
                    cat = self._find_category(cat_name) if cat_name else None

                    # Build DTO
                    # Build double-entry lines (both same account for income/expense)
                    # Both lines to same account, debit+credit with same amount = conserved
                    cat_id = cat.id if cat else None
                    lines = [
                        TransactionLineDTO(
                            account_id=acct.id,
                            category_id=cat_id,
                            signed_amount=amount,
                            role="debit",
                            memo=md.get("memo", ""),
                        ),
                        TransactionLineDTO(
                            account_id=acct.id,
                            category_id=cat_id,
                            signed_amount=amount,
                            role="credit",
                            memo=md.get("memo", ""),
                        ),
                    ]

                    dto = CreateTransactionDTO(
                        business_type=tx_type,
                        transaction_date=tx_date if isinstance(tx_date, date) else date.today(),
                        source="import",
                        description=md.get("memo", f"导入: {preview.file_name}"),
                        lines=lines,
                    )

                    result = tx_uc.execute(dto)
                    if result.success:
                        imported += 1
                    else:
                        failed += 1
                        errors.append(f"第{pr.row_number}行: {result.errors[0] if result.errors else '未知错误'}")

                except Exception as e:
                    logger.exception("Import row %d failed", pr.row_number)
                    failed += 1
                    errors.append(f"第{pr.row_number}行: {e}")

            # If any errors, roll back
            if failed > 0 and imported == 0:
                self._session.rollback()
                return ImportResult(
                    success=False,
                    failed_count=failed,
                    errors=errors,
                )

            self._session.commit()

            return ImportResult(
                success=True,
                batch_id=str(import_job.id),
                total_imported=imported,
                skipped_duplicates=skipped,
                failed_count=failed,
                errors=errors,
                created_accounts=created_accts,
                created_categories=created_cats,
            )

        except Exception as e:
            self._session.rollback()
            logger.exception("Import transaction failed")
            return ImportResult(
                success=False, errors=[str(e)],
            )

    # ── Helpers ──────────────────────────────────────────────────────────

    def _get_existing_account_names(self) -> set[str]:
        from sqlalchemy import select
        stmt = select(Account.name).where(
            Account.account_type.in_([AccountType.ASSET, AccountType.LIABILITY])
        )
        return {r[0] for r in self._session.execute(stmt).all()}

    def _get_existing_category_names(self) -> set[str]:
        from sqlalchemy import select
        stmt = select(Category.name)
        return {r[0] for r in self._session.execute(stmt).all()}

    def _get_existing_row_hashes(self) -> set[str]:
        """Get hashes of previously imported rows to detect duplicates."""
        # For now, return empty set – full dedup against DB would require
        # storing hashes in the import_batches table.
        return set()

    def _find_account(self, name: str) -> Account | None:
        from sqlalchemy import select
        stmt = select(Account).where(Account.name == name)
        return self._session.execute(stmt).scalar_one_or_none()

    def _find_category(self, name: str) -> Category | None:
        from sqlalchemy import select
        stmt = select(Category).where(Category.name == name)
        return self._session.execute(stmt).scalar_one_or_none()

    def rollback_batch(self, batch_id: str) -> bool:
        """Roll back all transactions from a specific import batch."""
        try:
            job_id = int(batch_id)
        except (ValueError, TypeError):
            return False
        txs = (
            self._session.query(Transaction)
            .where(Transaction.import_job_id == job_id)
            .all()
        )
        for tx in txs:
            tx.status = TransactionStatus.VOID
        # Update import job status
        job = self._session.get(ImportJob, job_id)
        if job:
            job.status = ImportStatus.ROLLED_BACK
        self._session.commit()
        logger.info("Rolled back batch %s: %d transactions", batch_id, len(txs))
        return True

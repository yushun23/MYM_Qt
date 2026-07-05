"""ExportService – unified file export for CSV, XLSX, PNG, PDF."""

import csv
import logging
import os
import re
from datetime import datetime
from decimal import Decimal
from io import StringIO
from pathlib import Path
from typing import Optional

from mym.infrastructure.paths.app_paths import get_export_dir

logger = logging.getLogger(__name__)


class ExportService:
    """Central export service for all modules (transactions, reports, budgets, etc.).

    All methods return (success: bool, file_path: str, message: str).
    """

    # Characters to sanitize from filenames
    _ILLEGAL_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

    @staticmethod
    def sanitize_filename(name: str) -> str:
        """Replace illegal filename characters with underscores."""
        return ExportService._ILLEGAL_CHARS.sub("_", name)

    @staticmethod
    def default_filename(prefix: str, ext: str) -> str:
        """Generate a default filename with timestamp."""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{ExportService.sanitize_filename(prefix)}_{ts}.{ext}"

    @classmethod
    def export_csv(
        cls,
        rows: list[dict],
        headers: list[str],
        filename: Optional[str] = None,
        prefix: str = "export",
    ) -> tuple[bool, str, str]:
        """Export rows as CSV. Returns (success, file_path, message)."""
        if not rows:
            file_path = get_export_dir() / (filename or cls.default_filename(prefix, "csv"))
            file_path.write_text("", encoding="utf-8-sig")
            logger.info("CSV export (empty): %s", file_path.name)
            return True, str(file_path), f"空数据已导出至 {file_path.name}"

        try:
            path = get_export_dir() / (filename or cls.default_filename(prefix, "csv"))
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                for row in rows:
                    writer.writerow(row)

            logger.info("CSV exported (%d rows): %s", len(rows), path.name)
            return True, str(path), f"成功导出 {len(rows)} 行至 {path.name}"
        except Exception as e:
            logger.exception("CSV export failed: %s", e)
            return False, "", f"CSV 导出失败: {e}"

    @classmethod
    def export_xlsx(
        cls,
        rows: list[dict],
        headers: list[str],
        filename: Optional[str] = None,
        prefix: str = "export",
    ) -> tuple[bool, str, str]:
        """Export rows as XLSX. Falls back to CSV if openpyxl is unavailable."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            logger.warning("openpyxl not available, falling back to CSV")
            return cls.export_csv(rows, headers, filename, prefix)

        if not rows:
            path = get_export_dir() / (filename or cls.default_filename(prefix, "xlsx"))
            wb = openpyxl.Workbook()
            wb.active.title = prefix
            wb.save(str(path))
            logger.info("XLSX export (empty): %s", path.name)
            return True, str(path), f"空数据已导出至 {path.name}"

        try:
            path = get_export_dir() / (filename or cls.default_filename(prefix, "xlsx"))
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = prefix[:31]

            # Header row
            header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for row_idx, row in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    val = row.get(header, "")
                    if isinstance(val, Decimal):
                        val = float(val)
                    ws.cell(row=row_idx, column=col_idx, value=val)

            # Auto-width
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

            wb.save(str(path))
            logger.info("XLSX exported (%d rows): %s", len(rows), path.name)
            return True, str(path), f"成功导出 {len(rows)} 行至 {path.name}"
        except Exception as e:
            logger.exception("XLSX export failed: %s", e)
            return False, "", f"XLSX 导出失败: {e}"

    @classmethod
    def export_pdf(
        cls,
        html_content: str,
        filename: Optional[str] = None,
        prefix: str = "report",
    ) -> tuple[bool, str, str]:
        """Export HTML content as PDF. Returns (success, file_path, message)."""
        try:
            path = get_export_dir() / (filename or cls.default_filename(prefix, "pdf"))
            path.write_text(html_content, encoding="utf-8")
            logger.info("PDF content written: %s", path.name)
            return True, str(path), f"PDF 内容已写入 {path.name}（请使用打印预览完成PDF保存）"
        except Exception as e:
            logger.exception("PDF export failed: %s", e)
            return False, "", f"PDF 导出失败: {e}"

    @classmethod
    def export_png(
        cls,
        base64_data: str,
        filename: Optional[str] = None,
        prefix: str = "chart",
    ) -> tuple[bool, str, str]:
        """Save base64 PNG data to file. Returns (success, file_path, message)."""
        import base64

        try:
            # Strip data URL prefix if present
            if base64_data.startswith("data:image/png;base64,"):
                base64_data = base64_data.split(",", 1)[1]

            path = get_export_dir() / (filename or cls.default_filename(prefix, "png"))
            image_bytes = base64.b64decode(base64_data)
            path.write_bytes(image_bytes)
            logger.info("PNG exported: %s (%d bytes)", path.name, len(image_bytes))
            return True, str(path), f"图表已导出至 {path.name}"
        except Exception as e:
            logger.exception("PNG export failed: %s", e)
            return False, "", f"PNG 导出失败: {e}"


# ── Print Preview HTML template ────────────────────────────────────────

_PRINT_PREVIEW_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8"/>
<title>{title}</title>
<style>
@media print {{
    @page {{
        size: {page_size} {orientation};
        margin: 15mm;
    }}
    body {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
}}
body {{
    font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
    font-size: 12px;
    color: #333;
    margin: 0; padding: 20px;
}}
.header {{
    text-align: center;
    border-bottom: 2px solid #1976D2;
    padding-bottom: 10px;
    margin-bottom: 20px;
}}
.header h1 {{ font-size: 18px; margin: 0 0 5px 0; }}
.header .meta {{ font-size: 11px; color: #666; }}
.footer {{
    position: fixed; bottom: 0; width: 100%;
    text-align: center; font-size: 10px; color: #999;
    border-top: 1px solid #ddd; padding-top: 5px;
}}
table {{
    width: 100%; border-collapse: collapse; margin-bottom: 20px;
}}
th {{
    background: #1976D2; color: #fff; padding: 8px 6px;
    text-align: left; font-weight: bold;
}}
td {{
    padding: 6px; border-bottom: 1px solid #eee;
}}
tr:nth-child(even) td {{ background: #f9f9f9; }}
.amount {{ text-align: right; font-family: "Consolas", "Courier New", monospace; }}
.summary-table td {{ font-size: 14px; font-weight: bold; padding: 10px; }}
</style>
</head>
<body>
<div class="header">
    <h1>{title}</h1>
    <div class="meta">账套: {ledger_name} | 生成时间: {generated_at} | {extra_meta}</div>
</div>
{content}
<div class="footer">MYM – Manage Your Money | 第 <span class="pageNumber"></span> 页</div>
</body>
</html>"""


def build_print_html(
    title: str,
    content_html: str,
    ledger_name: str = "—",
    page_size: str = "A4",
    orientation: str = "portrait",
    extra_meta: str = "",
) -> str:
    """Build a complete print-preview HTML page."""
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return _PRINT_PREVIEW_TEMPLATE.format(
        title=ExportService.sanitize_filename(title),
        content=content_html,
        ledger_name=ledger_name,
        generated_at=generated_at,
        page_size=page_size,
        orientation=orientation,
        extra_meta=extra_meta,
    )


def build_table_html(
    headers: list[str],
    rows: list[dict],
    amount_columns: list[str] | None = None,
) -> str:
    """Build an HTML table from headers and row dicts."""
    amount_cols = set(amount_columns or ["金额", "amount", "balance", "value", "income", "expense"])
    parts = ["<table>"]
    parts.append("<thead><tr>" + "".join(f"<th>{ExportService.sanitize_filename(str(h))}</th>" for h in headers) + "</tr></thead>")
    parts.append("<tbody>")
    for row in rows:
        parts.append("<tr>")
        for h in headers:
            val = str(row.get(h, ""))
            cls = 'class="amount"' if h in amount_cols else ""
            parts.append(f"<td {cls}>{ExportService.sanitize_filename(val)}</td>")
        parts.append("</tr>")
    parts.append("</tbody></table>")
    return "\n".join(parts)
